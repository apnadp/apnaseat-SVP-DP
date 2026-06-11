"""
apnaseat_bot.py
===============
Complete Selenium automation bot for ApnaSeat.in bus ticket booking.

Architecture
------------
ApnaSeatBot          – orchestrator; calls each step in order
  ├── BrowserManager – sets up Chrome with anti-detection options
  ├── SearchEngine   – fills route form & scrapes bus cards
  ├── BusRanker      – scores & picks the best bus
  ├── SeatSelector   – opens seat map & picks best seat
  └── PassengerForm  – fills passenger details from JSON
"""

from __future__ import annotations

import json
import logging
import os
import random
import sys
import time
import re
from dataclasses import dataclass, field
from datetime import datetime, timedelta
from pathlib import Path
from typing import List, Optional

from selenium import webdriver
from selenium.common.exceptions import (
    ElementClickInterceptedException,
    ElementNotInteractableException,
    NoSuchElementException,
    StaleElementReferenceException,
    TimeoutException,
)
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select, WebDriverWait

import config


# ─────────────────────────── logging ────────────────────────────

def setup_logging() -> logging.Logger:
    Path("logs").mkdir(exist_ok=True)
    fmt = "%(asctime)s [%(levelname)s] %(name)s — %(message)s"
    logging.basicConfig(
        level=logging.INFO,
        format=fmt,
        handlers=[
            logging.FileHandler(config.LOG_FILE, encoding="utf-8"),
            logging.FileHandler(Path("logs") / "application.log", encoding="utf-8"),
            logging.StreamHandler(open(sys.stdout.fileno(), mode='w', encoding='utf-8', closefd=False)),
        ],
    )
    return logging.getLogger("ApnaSeatBot")


log = setup_logging()


# ─────────────────────────── passenger directory ────────────────

_PASSENGER_DIRECTORY = [
    {"name": "Rajib Das",        "father_name": "Raju Das",          "id_number": "4123-5678-9012", "email": "rajibdas101@gmail.com",      "residence": "Aberdeen"},
    {"name": "Amit Roy",         "father_name": "Subhash Roy",       "id_number": "5234-6789-0123", "email": "amitroy102@gmail.com",       "residence": "Haddo"},
    {"name": "Suman Ghosh",      "father_name": "Pradip Ghosh",      "id_number": "6345-7890-1234", "email": "sumanghosh103@gmail.com",    "residence": "Junglighat"},
    {"name": "Rajesh Mondal",    "father_name": "Anil Mondal",       "id_number": "7456-8901-2345", "email": "rajeshmondal104@gmail.com",  "residence": "Dollygunj"},
    {"name": "Bikash Dey",       "father_name": "Shyamal Dey",       "id_number": "8567-9012-3456", "email": "bikashdey105@gmail.com",     "residence": "Garacharma"},
    {"name": "Sourav Paul",      "father_name": "Gopal Paul",        "id_number": "9678-0123-4567", "email": "souravpaul106@gmail.com",    "residence": "Shadipur"},
    {"name": "Debasis Sen",      "father_name": "Tapan Sen",         "id_number": "1789-1234-5678", "email": "debasissen107@gmail.com",    "residence": "Pahargaon"},
    {"name": "Arindam Roy",      "father_name": "Dilip Roy",         "id_number": "2890-2345-6789", "email": "arindamroy108@gmail.com",    "residence": "South Point"},
    {"name": "Sanjay Das",       "father_name": "Prakash Das",       "id_number": "3901-3456-7890", "email": "sanjaydas109@gmail.com",     "residence": "Minnie Bay"},
    {"name": "Ranjit Biswas",    "father_name": "Nirmal Biswas",     "id_number": "4012-4567-8901", "email": "ranjitbiswas110@gmail.com",  "residence": "Govinda Nagar"},
    {"name": "Manoj Dutta",      "father_name": "Tarun Dutta",       "id_number": "5123-5678-9013", "email": "manojdutta111@gmail.com",    "residence": "Vijaya Nagar"},
    {"name": "Prasenjit Pal",    "father_name": "Sukumar Pal",       "id_number": "6234-6789-0124", "email": "prasenjitpal112@gmail.com",  "residence": "Krishna Nagar"},
    {"name": "Kaushik Roy",      "father_name": "Kartik Roy",        "id_number": "7345-7890-1235", "email": "kaushikroy113@gmail.com",    "residence": "Radha Nagar"},
    {"name": "Subrata Das",      "father_name": "Bimal Das",         "id_number": "8456-8901-2346", "email": "subratadas114@gmail.com",    "residence": "Bharatpur"},
    {"name": "Abhijit Sen",      "father_name": "Haripada Sen",      "id_number": "9567-9012-3457", "email": "abhijitsen115@gmail.com",    "residence": "Neil Kendra"},
    {"name": "Tapan Dey",        "father_name": "Nikhil Dey",        "id_number": "1678-0123-4568", "email": "tapandey116@gmail.com",      "residence": "Lakshmanpur"},
    {"name": "Dipankar Roy",     "father_name": "Suresh Roy",        "id_number": "2789-1234-5679", "email": "dipankarroy117@gmail.com",   "residence": "Ram Nagar"},
    {"name": "Joydeep Ghosh",    "father_name": "Prabir Ghosh",      "id_number": "3890-2345-6780", "email": "joydeepghosh118@gmail.com",  "residence": "School Line"},
    {"name": "Sujit Paul",       "father_name": "Manoranjan Paul",   "id_number": "4901-3456-7891", "email": "sujitpaul119@gmail.com",     "residence": "Brooksabad"},
    {"name": "Anup Das",         "father_name": "Gour Das",          "id_number": "5012-4567-8902", "email": "anupdas120@gmail.com",       "residence": "Rangachang"},
    {"name": "Prabir Roy",       "father_name": "Ajit Roy",          "id_number": "6123-5678-9014", "email": "prabirroy121@gmail.com",     "residence": "Chidiya Tapu"},
    {"name": "Kunal Ghosh",      "father_name": "Dinesh Ghosh",      "id_number": "7234-6789-0125", "email": "kunalghosh122@gmail.com",    "residence": "Calicut"},
    {"name": "Sandeep Dey",      "father_name": "Umesh Dey",         "id_number": "8345-7890-1236", "email": "sandeepdey123@gmail.com",    "residence": "Sippighat"},
    {"name": "Ashok Das",        "father_name": "Rabindra Das",      "id_number": "9456-8901-2347", "email": "ashokdas124@gmail.com",      "residence": "Taylorabad"},
    {"name": "Nayan Roy",        "father_name": "Keshab Roy",        "id_number": "1567-9012-3458", "email": "nayanroy125@gmail.com",      "residence": "Buniyadabad"},
    {"name": "Pratik Sen",       "father_name": "Madan Sen",         "id_number": "2678-0123-4569", "email": "pratiksen126@gmail.com",     "residence": "Austinabad"},
    {"name": "Arup Das",         "father_name": "Narayan Das",       "id_number": "3789-1234-5670", "email": "arupdas127@gmail.com",       "residence": "Nayagaon"},
    {"name": "Supriyo Roy",      "father_name": "Pankaj Roy",        "id_number": "4890-2345-6781", "email": "supriyoroy128@gmail.com",    "residence": "Dudh Line"},
    {"name": "Gautam Paul",      "father_name": "Ratan Paul",        "id_number": "5901-3456-7892", "email": "gautampaul129@gmail.com",    "residence": "Lambaline"},
    {"name": "Sandip Biswas",    "father_name": "Uttam Biswas",      "id_number": "6012-4567-8903", "email": "sandipbiswas130@gmail.com",  "residence": "Goodwill Estate"},
    {"name": "Alok Dey",         "father_name": "Bijoy Dey",         "id_number": "7123-5678-9015", "email": "alokdey131@gmail.com",       "residence": "Phoenix Bay"},
    {"name": "Bappa Roy",        "father_name": "Mihir Roy",         "id_number": "8234-6789-0126", "email": "bapparoy132@gmail.com",      "residence": "Carbyns' Cove"},
    {"name": "Tapas Das",        "father_name": "Samar Das",         "id_number": "9345-7890-1237", "email": "tapasdas133@gmail.com",      "residence": "Aberdeen Bazar"},
    {"name": "Sajal Sen",        "father_name": "Hemanta Sen",       "id_number": "1456-8901-2348", "email": "sajalsen134@gmail.com",      "residence": "Prothrapur"},
    {"name": "Biswajit Roy",     "father_name": "Paresh Roy",        "id_number": "2567-9012-3459", "email": "biswajitroy135@gmail.com",   "residence": "Beodnabad"},
    {"name": "Rakesh Dey",       "father_name": "Chandan Dey",       "id_number": "3678-0123-4570", "email": "rakeshdey136@gmail.com",     "residence": "Brich Gunj"},
    {"name": "Anirban Das",      "father_name": "Monoranjan Das",    "id_number": "4789-1234-5671", "email": "anirbandas137@gmail.com",    "residence": "Bimblitan"},
    {"name": "Soumen Roy",       "father_name": "Asit Roy",          "id_number": "5890-2345-6782", "email": "soumenroy138@gmail.com",     "residence": "Rutland"},
    {"name": "Debjit Sen",       "father_name": "Tarak Sen",         "id_number": "6901-3456-7893", "email": "debjitsen139@gmail.com",     "residence": "Govinda Nagar"},
    {"name": "Kaustav Das",      "father_name": "Biresh Das",        "id_number": "7012-4567-8904", "email": "kaustavdas140@gmail.com",    "residence": "Krishna Nagar"},
    {"name": "Arnab Roy",        "father_name": "Tapan Roy",         "id_number": "8123-5678-9016", "email": "arnabroy141@gmail.com",      "residence": "Vijaya Nagar"},
    {"name": "Shubham Ghosh",    "father_name": "Dilip Ghosh",       "id_number": "9234-6789-0127", "email": "shubhamghosh142@gmail.com",  "residence": "Bharatpur"},
    {"name": "Rohan Das",        "father_name": "Haran Das",         "id_number": "1345-7890-1238", "email": "rohandas143@gmail.com",      "residence": "South Point"},
    {"name": "Pritam Roy",       "father_name": "Sukhen Roy",        "id_number": "2456-8901-2349", "email": "pritamroy144@gmail.com",     "residence": "Pahargaon"},
    {"name": "Debabrata Sen",    "father_name": "Kalyan Sen",        "id_number": "3567-9012-3460", "email": "debabratasen145@gmail.com",  "residence": "Minnie Bay"},
    {"name": "Tanmoy Das",       "father_name": "Nityananda Das",    "id_number": "4678-0123-4571", "email": "tanmoydas146@gmail.com",     "residence": "Shadipur"},
    {"name": "Joy Roy",          "father_name": "Nimai Roy",         "id_number": "5789-1234-5672", "email": "joyroy147@gmail.com",        "residence": "Junglighat"},
    {"name": "Rupam Dey",        "father_name": "Goutam Dey",        "id_number": "6890-2345-6783", "email": "rupamdey148@gmail.com",      "residence": "Dollygunj"},
    {"name": "Abir Sen",         "father_name": "Shankar Sen",       "id_number": "7901-3456-7894", "email": "abirsen149@gmail.com",       "residence": "Garacharma"},
    {"name": "Sayan Das",        "father_name": "Pradip Das",        "id_number": "8012-4567-8905", "email": "sayandas150@gmail.com",      "residence": "Chidiya Tapu"},
    {"name":"Sayan Das","father_name":"Pradip Das","id_number":"8012-4567-8905","email":"sayandas150@gmail.com","residence":"Chidiya Tapu"},
    {"name":"Subhajit Roy","father_name":"Sukumar Roy","id_number":"8012-4567-8906","email":"subhajitroy151@gmail.com","residence":"Chidiya Tapu"},
    {"name":"Anirban Das","father_name":"Prabir Das","id_number":"8012-4567-8907","email":"anirbandas152@gmail.com","residence":"Chidiya Tapu"},
    {"name":"Sourav Ghosh","father_name":"Bimal Ghosh","id_number":"8012-4567-8908","email":"souravghosh153@gmail.com","residence":"Chidiya Tapu"},
    {"name":"Arindam Paul","father_name":"Biswajit Paul","id_number":"8012-4567-8909","email":"arindampaul154@gmail.com","residence":"Chidiya Tapu"},
    {"name":"Debasis Dey","father_name":"Dipak Dey","id_number":"8012-4567-8910","email":"debasisdey155@gmail.com","residence":"Chidiya Tapu"},
    {"name":"Abhijit Mondal","father_name":"Tarun Mondal","id_number":"8012-4567-8911","email":"abhijitmondal156@gmail.com","residence":"Chidiya Tapu"},
    {"name":"Ritam Chakraborty","father_name":"Swapan Chakraborty","id_number":"8012-4567-8912","email":"ritam157@gmail.com","residence":"Chidiya Tapu"},
    {"name":"Kaushik Banerjee","father_name":"Ashok Banerjee","id_number":"8012-4567-8913","email":"kaushik158@gmail.com","residence":"Chidiya Tapu"},
    {"name":"Sagnik Dutta","father_name":"Partha Dutta","id_number":"8012-4567-8914","email":"sagnik159@gmail.com","residence":"Chidiya Tapu"},
    {"name":"Amit Saha","father_name":"Nirmal Saha","id_number":"8012-4567-8915","email":"amit160@gmail.com","residence":"Chidiya Tapu"},
    {"name":"Rajdeep Sen","father_name":"Prasenjit Sen","id_number":"8012-4567-8916","email":"rajdeep161@gmail.com","residence":"Chidiya Tapu"},
    {"name":"Soumik Biswas","father_name":"Goutam Biswas","id_number":"8012-4567-8917","email":"soumik162@gmail.com","residence":"Chidiya Tapu"},
    {"name":"Tanmoy Pal","father_name":"Madan Pal","id_number":"8012-4567-8918","email":"tanmoy163@gmail.com","residence":"Chidiya Tapu"},
    {"name":"Ayan Chatterjee","father_name":"Anup Chatterjee","id_number":"8012-4567-8919","email":"ayan164@gmail.com","residence":"Chidiya Tapu"},
    {"name":"Arnab Mukherjee","father_name":"Tapan Mukherjee","id_number":"8012-4567-8920","email":"arnab165@gmail.com","residence":"Chidiya Tapu"},
    {"name":"Niladri Sarkar","father_name":"Dilip Sarkar","id_number":"8012-4567-8921","email":"niladri166@gmail.com","residence":"Chidiya Tapu"},
    {"name":"Suman Kundu","father_name":"Ranjit Kundu","id_number":"8012-4567-8922","email":"suman167@gmail.com","residence":"Chidiya Tapu"},
    {"name":"Bikram Das","father_name":"Shyamal Das","id_number":"8012-4567-8923","email":"bikram168@gmail.com","residence":"Chidiya Tapu"},
    {"name":"Rohit Roy","father_name":"Mihir Roy","id_number":"8012-4567-8924","email":"rohit169@gmail.com","residence":"Chidiya Tapu"},
    {"name":"Sanjib Dutta","father_name":"Anil Dutta","id_number":"8012-4567-8925","email":"sanjib170@gmail.com","residence":"Chidiya Tapu"},
    {"name":"Prasenjit Pal","father_name":"Bapi Pal","id_number":"8012-4567-8926","email":"prasenjit171@gmail.com","residence":"Chidiya Tapu"},
    {"name":"Debjit Roy","father_name":"Kamal Roy","id_number":"8012-4567-8927","email":"debjit172@gmail.com","residence":"Chidiya Tapu"},
    {"name":"Saptarshi Das","father_name":"Subal Das","id_number":"8012-4567-8928","email":"saptarshi173@gmail.com","residence":"Chidiya Tapu"},
    {"name":"Joydeep Sen","father_name":"Asim Sen","id_number":"8012-4567-8929","email":"joydeep174@gmail.com","residence":"Chidiya Tapu"},
    {"name":"Rupam Ghosh","father_name":"Ratan Ghosh","id_number":"8012-4567-8930","email":"rupam175@gmail.com","residence":"Chidiya Tapu"},
    {"name":"Sandip Paul","father_name":"Gobinda Paul","id_number":"8012-4567-8931","email":"sandip176@gmail.com","residence":"Chidiya Tapu"},
    {"name":"Koushik Dey","father_name":"Narayan Dey","id_number":"8012-4567-8932","email":"koushik177@gmail.com","residence":"Chidiya Tapu"},
    {"name":"Partha Roy","father_name":"Bhuban Roy","id_number":"8012-4567-8933","email":"partha178@gmail.com","residence":"Chidiya Tapu"},
    {"name":"Indranil Das","father_name":"Sunil Das","id_number":"8012-4567-8934","email":"indranil179@gmail.com","residence":"Chidiya Tapu"},
    {"name":"Souradeep Sen","father_name":"Pradip Sen","id_number":"8012-4567-8935","email":"souradeep180@gmail.com","residence":"Chidiya Tapu"},
    {"name":"Anupam Paul","father_name":"Tarak Paul","id_number":"8012-4567-8936","email":"anupam181@gmail.com","residence":"Chidiya Tapu"},
    {"name":"Biswajit Ghosh","father_name":"Sajal Ghosh","id_number":"8012-4567-8937","email":"biswajit182@gmail.com","residence":"Chidiya Tapu"},
    {"name":"Debabrata Roy","father_name":"Milan Roy","id_number":"8012-4567-8938","email":"debabrata183@gmail.com","residence":"Chidiya Tapu"},
    {"name":"Sourav Mondal","father_name":"Biren Mondal","id_number":"8012-4567-8939","email":"sourav184@gmail.com","residence":"Chidiya Tapu"},
    {"name":"Arup Das","father_name":"Tushar Das","id_number":"8012-4567-8940","email":"arup185@gmail.com","residence":"Chidiya Tapu"},
    {"name":"Tirthankar Sen","father_name":"Ajit Sen","id_number":"8012-4567-8941","email":"tirthankar186@gmail.com","residence":"Chidiya Tapu"},
    {"name":"Subrata Dey","father_name":"Madan Dey","id_number":"8012-4567-8942","email":"subrata187@gmail.com","residence":"Chidiya Tapu"},
    {"name":"Prabir Das","father_name":"Nitai Das","id_number":"8012-4567-8943","email":"prabir188@gmail.com","residence":"Chidiya Tapu"},
    {"name":"Sukanta Roy","father_name":"Shankar Roy","id_number":"8012-4567-8944","email":"sukanta189@gmail.com","residence":"Chidiya Tapu"},
    {"name":"Amlan Ghosh","father_name":"Prafulla Ghosh","id_number":"8012-4567-8945","email":"amlan190@gmail.com","residence":"Chidiya Tapu"},
    {"name":"Ranjan Dutta","father_name":"Bholanath Dutta","id_number":"8012-4567-8946","email":"ranjan191@gmail.com","residence":"Chidiya Tapu"},
    {"name":"Sujoy Pal","father_name":"Kalipada Pal","id_number":"8012-4567-8947","email":"sujoy192@gmail.com","residence":"Chidiya Tapu"},
    {"name":"Pritam Das","father_name":"Haripada Das","id_number":"8012-4567-8948","email":"pritam193@gmail.com","residence":"Chidiya Tapu"},
    {"name":"Kaustav Roy","father_name":"Dinesh Roy","id_number":"8012-4567-8949","email":"kaustav194@gmail.com","residence":"Chidiya Tapu"},
    {"name":"Soham Ghosh","father_name":"Keshab Ghosh","id_number":"8012-4567-8950","email":"soham195@gmail.com","residence":"Chidiya Tapu"},
    {"name":"Abir Das","father_name":"Mihir Das","id_number":"8012-4567-8951","email":"abir196@gmail.com","residence":"Chidiya Tapu"},
    {"name":"Ritwick Sen","father_name":"Pankaj Sen","id_number":"8012-4567-8952","email":"ritwick197@gmail.com","residence":"Chidiya Tapu"},
    {"name":"Mainak Roy","father_name":"Amal Roy","id_number":"8012-4567-8953","email":"mainak198@gmail.com","residence":"Chidiya Tapu"},
    {"name":"Aritra Das","father_name":"Sankar Das","id_number":"8012-4567-8954","email":"aritra199@gmail.com","residence":"Chidiya Tapu"},
    
{"name":"Arun Kumar","father_name":"Ramasamy","id_number":"8012-4567-9001","email":"arunkumar001@gmail.com","residence":"Port Blair"},
{"name":"Karthik Raj","father_name":"Muthukumar","id_number":"8012-4567-9002","email":"karthikraj002@gmail.com","residence":"Port Blair"},
{"name":"Suresh Babu","father_name":"Perumal","id_number":"8012-4567-9003","email":"sureshbabu003@gmail.com","residence":"Port Blair"},
{"name":"Vignesh Kumar","father_name":"Subramanian","id_number":"8012-4567-9004","email":"vignesh004@gmail.com","residence":"Port Blair"},
{"name":"Praveen Raj","father_name":"Selvaraj","id_number":"8012-4567-9005","email":"praveen005@gmail.com","residence":"Port Blair"},
{"name":"Dinesh Kumar","father_name":"Murugan","id_number":"8012-4567-9006","email":"dinesh006@gmail.com","residence":"Port Blair"},
{"name":"Saravanan","father_name":"Ganesan","id_number":"8012-4567-9007","email":"saravanan007@gmail.com","residence":"Port Blair"},
{"name":"Ramesh Kumar","father_name":"Ranganathan","id_number":"8012-4567-9008","email":"ramesh008@gmail.com","residence":"Port Blair"},
{"name":"Hari Prasad","father_name":"Krishnan","id_number":"8012-4567-9009","email":"hari009@gmail.com","residence":"Port Blair"},
{"name":"Sathish Kumar","father_name":"Velmurugan","id_number":"8012-4567-9010","email":"sathish010@gmail.com","residence":"Port Blair"},
{"name":"Manikandan","father_name":"Rajendran","id_number":"8012-4567-9011","email":"manikandan011@gmail.com","residence":"Port Blair"},
{"name":"Bharath Kumar","father_name":"Shanmugam","id_number":"8012-4567-9012","email":"bharath012@gmail.com","residence":"Port Blair"},
{"name":"Senthil Kumar","father_name":"Natarajan","id_number":"8012-4567-9013","email":"senthil013@gmail.com","residence":"Port Blair"},
{"name":"Vinoth Kumar","father_name":"Balasubramanian","id_number":"8012-4567-9014","email":"vinoth014@gmail.com","residence":"Port Blair"},
{"name":"Madhan Kumar","father_name":"Sivakumar","id_number":"8012-4567-9015","email":"madhan015@gmail.com","residence":"Port Blair"},
{"name":"Aravind Raj","father_name":"Kandasamy","id_number":"8012-4567-9016","email":"aravind016@gmail.com","residence":"Port Blair"},
{"name":"Jagan Mohan","father_name":"Duraisamy","id_number":"8012-4567-9017","email":"jagan017@gmail.com","residence":"Port Blair"},
{"name":"Naveen Kumar","father_name":"Pandian","id_number":"8012-4567-9018","email":"naveen018@gmail.com","residence":"Port Blair"},
{"name":"Gokul Raj","father_name":"Sundaram","id_number":"8012-4567-9019","email":"gokul019@gmail.com","residence":"Port Blair"},
{"name":"Prakash Kumar","father_name":"Ravi","id_number":"8012-4567-9020","email":"prakash020@gmail.com","residence":"Port Blair"},
{"name":"Muthu Kumar","father_name":"Palanisamy","id_number":"8012-4567-9021","email":"muthu021@gmail.com","residence":"Port Blair"},
{"name":"Ajith Kumar","father_name":"Thangaraj","id_number":"8012-4567-9022","email":"ajith022@gmail.com","residence":"Port Blair"},
{"name":"Kamal Raj","father_name":"Arumugam","id_number":"8012-4567-9023","email":"kamal023@gmail.com","residence":"Port Blair"},
{"name":"Siva Kumar","father_name":"Chandrasekar","id_number":"8012-4567-9024","email":"siva024@gmail.com","residence":"Port Blair"},
{"name":"Ranjith Kumar","father_name":"Manoharan","id_number":"8012-4567-9025","email":"ranjith025@gmail.com","residence":"Port Blair"},
{"name":"Anil Kumar","father_name":"Raghavan","id_number":"8012-4567-9101","email":"anilkumar101@gmail.com","residence":"Port Blair"},
{"name":"Suresh Nair","father_name":"Madhavan","id_number":"8012-4567-9102","email":"suresh102@gmail.com","residence":"Port Blair"},
{"name":"Rajesh Kumar","father_name":"Gopalakrishnan","id_number":"8012-4567-9103","email":"rajesh103@gmail.com","residence":"Port Blair"},
{"name":"Pradeep Kumar","father_name":"Narayanan","id_number":"8012-4567-9104","email":"pradeep104@gmail.com","residence":"Port Blair"},
{"name":"Sunil Kumar","father_name":"Krishnan","id_number":"8012-4567-9105","email":"sunil105@gmail.com","residence":"Port Blair"},
{"name":"Ajith Kumar","father_name":"Sasidharan","id_number":"8012-4567-9106","email":"ajith106@gmail.com","residence":"Port Blair"},
{"name":"Vijay Kumar","father_name":"Ramakrishnan","id_number":"8012-4567-9107","email":"vijay107@gmail.com","residence":"Port Blair"},
{"name":"Biju Thomas","father_name":"Thomas","id_number":"8012-4567-9108","email":"biju108@gmail.com","residence":"Port Blair"},
{"name":"Shyam Kumar","father_name":"Vasudevan","id_number":"8012-4567-9109","email":"shyam109@gmail.com","residence":"Port Blair"},
{"name":"Arun Nair","father_name":"Soman","id_number":"8012-4567-9110","email":"arun110@gmail.com","residence":"Port Blair"},
{"name":"Ratheesh Kumar","father_name":"Balakrishnan","id_number":"8012-4567-9111","email":"ratheesh111@gmail.com","residence":"Port Blair"},
{"name":"Jithin Raj","father_name":"Mohanan","id_number":"8012-4567-9112","email":"jithin112@gmail.com","residence":"Port Blair"},
{"name":"Nithin Kumar","father_name":"Raveendran","id_number":"8012-4567-9113","email":"nithin113@gmail.com","residence":"Port Blair"},
{"name":"Sajith Kumar","father_name":"Thankappan","id_number":"8012-4567-9114","email":"sajith114@gmail.com","residence":"Port Blair"},
{"name":"Manoj Kumar","father_name":"Damodaran","id_number":"8012-4567-9115","email":"manoj115@gmail.com","residence":"Port Blair"},
{"name":"Sandeep Kumar","father_name":"Haridas","id_number":"8012-4567-9116","email":"sandeep116@gmail.com","residence":"Port Blair"},
{"name":"Anoop Kumar","father_name":"Rajan","id_number":"8012-4567-9117","email":"anoop117@gmail.com","residence":"Port Blair"},
{"name":"Deepak Nair","father_name":"Chandran","id_number":"8012-4567-9118","email":"deepak118@gmail.com","residence":"Port Blair"},
{"name":"Rohith Kumar","father_name":"Jayakumar","id_number":"8012-4567-9119","email":"rohith119@gmail.com","residence":"Port Blair"},
{"name":"Vinod Kumar","father_name":"Prabhakaran","id_number":"8012-4567-9120","email":"vinod120@gmail.com","residence":"Port Blair"},
{"name":"Sijin Thomas","father_name":"Joseph","id_number":"8012-4567-9121","email":"sijin121@gmail.com","residence":"Port Blair"},
{"name":"Akhil Kumar","father_name":"Radhakrishnan","id_number":"8012-4567-9122","email":"akhil122@gmail.com","residence":"Port Blair"},
{"name":"Rahul Nair","father_name":"Sukumaran","id_number":"8012-4567-9123","email":"rahul123@gmail.com","residence":"Port Blair"},
{"name":"Jaison Mathew","father_name":"Mathew","id_number":"8012-4567-9124","email":"jaison124@gmail.com","residence":"Port Blair"},
{"name":"Bibin Joseph","father_name":"Joseph","id_number":"8012-4567-9125","email":"bibin125@gmail.com","residence":"Port Blair"}
]


# ─────────────────────────── data models ────────────────────────

@dataclass
class BusInfo:
    index: int                    # position in the search results
    name: str = ""
    departure: str = ""           # "HH:MM"
    arrival: str = ""
    fare: float = 9999.0
    available_seats: int = 0
    bus_type: str = ""
    score: float = 0.0
    element_ref: object = field(default=None, repr=False)  # Selenium element


@dataclass
class Passenger:
    name: str
    age: str
    gender: str
    phone: str
    email: str
    id_type: str
    id_number: str
    father_name: str = ""
    residence: str = ""

    @classmethod
    def from_json(cls, path: str) -> "Passenger":
        with open(path, encoding="utf-8") as f:
            d = json.load(f)
        return cls(**d)


def generate_passenger() -> Passenger:
    """Pick a random entry from the passenger directory and return a Passenger."""
    rec   = random.choice(_PASSENGER_DIRECTORY)
    phone = str(random.randint(6, 9)) + "".join(str(random.randint(0, 9)) for _ in range(9))
    return Passenger(
        name=rec["name"],
        age=str(random.randint(18, 60)),
        gender="Male",
        phone=phone,
        email=rec["email"],
        id_type="Aadhaar",
        id_number=rec["id_number"],
        father_name=rec["father_name"],
        residence=rec["residence"],
    )


def get_date_range() -> list:
    """Return journey dates as datetime objects: tomorrow through the next N days."""
    start = getattr(config, "BOOKING_START_OFFSET", 1)
    count = getattr(config, "BOOKING_DAY_COUNT", 4)
    today = datetime.now()
    dates = [today + timedelta(days=start + i) for i in range(count)]
    log.info(
        "📅 Journey date range: %s → %s (%d days)",
        dates[0].strftime("%d-%m-%Y"), dates[-1].strftime("%d-%m-%Y"), count,
    )
    return dates


# ─────────────────────────── browser ────────────────────────────

class BrowserManager:
    """Creates a stealthy, production-ready Chrome session."""

    @staticmethod
    def build() -> webdriver.Chrome:
        opts = Options()

        # ── anti-detection ──────────────────────────────────────
        opts.add_argument("--disable-blink-features=AutomationControlled")
        opts.add_experimental_option("excludeSwitches", ["enable-automation"])
        opts.add_experimental_option("useAutomationExtension", False)

        # ── stability ───────────────────────────────────────────
        import tempfile, atexit, shutil
        tmp_profile = tempfile.mkdtemp(prefix="chrome_bot_")
        atexit.register(shutil.rmtree, tmp_profile, True)
        opts.add_argument(f"--user-data-dir={tmp_profile}")
        opts.add_argument("--no-first-run")
        opts.add_argument("--no-default-browser-check")
        opts.add_argument("--disable-session-crashed-bubble")
        opts.add_argument("--disable-infobars")
        opts.add_argument("--window-size=1400,900")
        opts.add_argument("--disable-notifications")
        opts.add_experimental_option("prefs", {"profile.default_content_setting_values.notifications": 2})
        opts.add_argument("--start-maximized")
        opts.add_argument(
            "--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/124.0.0.0 Safari/537.36"
        )

        if config.HEADLESS:
            opts.add_argument("--headless=new")

        driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=opts)

        # Patch navigator.webdriver → undefined
        driver.execute_script(
            "Object.defineProperty(navigator, 'webdriver', {get: () => undefined})"
        )
        driver.implicitly_wait(3)
        return driver


# ─────────────────────────── helpers ────────────────────────────

class BotBase:
    def __init__(self, driver: webdriver.Chrome):
        self.driver = driver
        self.wait = WebDriverWait(driver, config.ELEMENT_WAIT_S)

    # ── waits ────────────────────────────────────────────────────

    def wait_for(self, by: str, value: str, timeout: int = None):
        t = timeout or config.ELEMENT_WAIT_S
        return WebDriverWait(self.driver, t).until(
            EC.presence_of_element_located((by, value))
        )

    def wait_visible(self, by: str, value: str, timeout: int = None):
        t = timeout or config.ELEMENT_WAIT_S
        return WebDriverWait(self.driver, t).until(
            EC.visibility_of_element_located((by, value))
        )

    def wait_clickable(self, by: str, value: str, timeout: int = None):
        t = timeout or config.ELEMENT_WAIT_S
        return WebDriverWait(self.driver, t).until(
            EC.element_to_be_clickable((by, value))
        )

    # ── safe interactions ────────────────────────────────────────

    def safe_click(self, element, retries: int = 3):
        for attempt in range(retries):
            try:
                self.driver.execute_script(
                    "arguments[0].scrollIntoView({block:'center'});", element
                )
                time.sleep(0.3)
                element.click()
                return
            except (ElementClickInterceptedException, StaleElementReferenceException):
                if attempt == retries - 1:
                    # JS click as last resort
                    self.driver.execute_script("arguments[0].click();", element)
                time.sleep(0.5)

    def safe_type(self, element, text: str):
        element.clear()
        time.sleep(0.2)
        element.send_keys(text)

    def js_set_value(self, element, value: str):
        self.driver.execute_script(
            "arguments[0].value = arguments[1];"
            "arguments[0].dispatchEvent(new Event('input',{bubbles:true}));"
            "arguments[0].dispatchEvent(new Event('change',{bubbles:true}));",
            element,
            value,
        )

    # ── screenshot ───────────────────────────────────────────────

    def screenshot(self, name: str):
        return  # screenshots disabled (Change 17)

    # ── text extraction ──────────────────────────────────────────

    @staticmethod
    def extract_number(text: str) -> float:
        """Pull the first number (incl. decimals) out of a string."""
        m = re.search(r"[\d,]+(?:\.\d+)?", text.replace(",", ""))
        return float(m.group().replace(",", "")) if m else 0.0

    @staticmethod
    def time_to_minutes(t: str) -> int:
        """'14:30' or '04:15AM' → minutes since midnight."""
        try:
            t = t.strip()
            pm = "PM" in t.upper()
            t = t.upper().replace("AM", "").replace("PM", "").strip()
            h, m = map(int, t.split(":"))
            if pm and h != 12:
                h += 12
            elif not pm and h == 12:
                h = 0
            return h * 60 + m
        except Exception:
            return 9999


# ─────────────────────────── Step 1 & 2: open + fill form ───────

class SearchEngine(BotBase):

    BASE_URL = "https://apnaseat.in/"

    def open_website(self):
        log.info("🌐 Opening %s", self.BASE_URL)
        self.driver.get(self.BASE_URL)
        # Wait for the body to be present
        self.wait_for(By.TAG_NAME, "body", timeout=config.PAGE_TIMEOUT_S)
        time.sleep(2)
        self.screenshot("01_homepage")
        log.info("✅ Website loaded")

    def fill_journey_form(
        self,
        source: str,
        destination: str,
        travel_date: str,
    ):
        log.info(
            "📋 Filling form — from: %s  to: %s  date: %s",
            source,
            destination,
            travel_date,
        )

        # ── Try quick-link buttons first (e.g. "SVP to DP") ─────
        if self._click_quick_route(source, destination):
            log.info("   ↳ Used quick-link button")
        else:
            # ── Source ───────────────────────────────────────────
            self._fill_route_field("source", source)
            time.sleep(1)
            # ── Destination ──────────────────────────────────────
            self._fill_route_field("destination", destination)
            time.sleep(1)

        # ── Date ─────────────────────────────────────────────────
        self._fill_date(travel_date)
        time.sleep(0.5)

        self.screenshot("02_form_filled")

    def _click_quick_route(self, source: str, destination: str) -> bool:
        """Click a quick-route button like 'SVP to DP' if one exists."""
        needle = f"{source} to {destination}".lower()
        try:
            buttons = self.driver.find_elements(By.CSS_SELECTOR, "button, a, span")
            for btn in buttons:
                if btn.text.strip().lower() == needle and btn.is_displayed():
                    self.safe_click(btn)
                    time.sleep(1)
                    return True
        except Exception:
            pass
        return False

    def _fill_route_field(self, field_role: str, value: str):
        """Fill a source/destination field that may be a <select> or autocomplete input."""
        # ── Try <select> first ───────────────────────────────────
        select_selectors = [
            f"select[name*='{field_role}' i]",
            f"select[id*='{field_role}' i]",
            f"select[class*='{field_role}' i]",
        ]
        if field_role == "source":
            select_selectors += ["select[name*='from' i]", "select[id*='from' i]", "select:nth-of-type(1)"]
        else:
            select_selectors += ["select[name*='to' i]", "select[id*='to' i]", "select:nth-of-type(2)"]

        for sel in select_selectors:
            try:
                el = self.driver.find_element(By.CSS_SELECTOR, sel)
                if el.is_displayed():
                    s = Select(el)
                    try:
                        s.select_by_visible_text(value)
                    except Exception:
                        s.select_by_value(value)
                    log.info("   ↳ %s = '%s' (select element)", field_role, value)
                    return
            except NoSuchElementException:
                pass

        # ── Fall back to autocomplete text input ─────────────────
        input_selectors = [
            f"input[placeholder*='{field_role}' i]",
            f"input[name*='{field_role}' i]",
            f"input[id*='{field_role}' i]",
        ]
        if field_role == "source":
            input_selectors += ["input[placeholder*='from' i]", "input[placeholder*='boarding' i]"]
        else:
            input_selectors += ["input[placeholder*='to' i]", "input[placeholder*='drop' i]"]

        inp = self._find_first(input_selectors)
        if inp is None:
            log.warning("⚠️  Could not locate %s field", field_role)
            return

        self.safe_click(inp)
        inp.clear()
        inp.send_keys(value)
        time.sleep(1.5)

        for sel in ["[class*='suggestion']", "[class*='option']", "[class*='item']", ".dropdown-item"]:
            items = self.driver.find_elements(By.CSS_SELECTOR, sel)
            if items:
                self.safe_click(items[0])
                log.info("   ↳ %s = '%s' (autocomplete)", field_role, value)
                return

        inp.send_keys(Keys.TAB)
        log.info("   ↳ %s = '%s' (typed)", field_role, value)

    def _fill_date(self, travel_date: str):
        """Fill the travel date input (handles date pickers & plain inputs)."""
        date_selectors = [
            "input[type='date']",
            "input[placeholder*='date' i]",
            "input[name*='date' i]",
            "input[id*='date' i]",
            "input[class*='date' i]",
        ]
        inp = self._find_first(date_selectors)
        if inp is None:
            log.warning("⚠️  Date field not found — skipping date fill")
            return

        # Try JS set (works for most date pickers)
        self.js_set_value(inp, travel_date)
        time.sleep(0.3)

        # Fallback: type character by character
        if not inp.get_attribute("value"):
            self.safe_click(inp)
            inp.send_keys(travel_date)

        log.info("   ↳ date = '%s'", travel_date)

    def _find_first(self, selectors: list):
        for sel in selectors:
            try:
                if sel.startswith("(//") or sel.startswith("//"):
                    el = self.driver.find_element(By.XPATH, sel)
                else:
                    el = self.driver.find_element(By.CSS_SELECTOR, sel)
                if el.is_displayed():
                    return el
            except NoSuchElementException:
                pass
        return None

    # ── Step 3: search ───────────────────────────────────────────

    def click_search(self):
        log.info("🔍 Clicking Search button …")
        btn_selectors = [
            "button[type='submit']",
            "input[type='submit']",
            "button[class*='search' i]",
            "a[class*='search' i]",
            "[class*='search-btn' i]",
            "button",
        ]
        for sel in btn_selectors:
            btns = self.driver.find_elements(By.CSS_SELECTOR, sel)
            for btn in btns:
                txt = btn.text.strip().lower()
                if any(k in txt for k in ("search", "find", "go", "book")):
                    log.info("   ↳ Clicked: '%s'", txt)
                    self.safe_click(btn)
                    return

        # Fallback: submit the form via JS
        self.driver.execute_script(
            "document.querySelector('form') && document.querySelector('form').submit()"
        )

    def wait_for_results(self):
        log.info("⏳ Waiting for bus results …")
        time.sleep(3)
        # ApnaSeat shows "Quick Book N" buttons when results load
        try:
            WebDriverWait(self.driver, config.PAGE_TIMEOUT_S).until(
                EC.presence_of_element_located(
                    (By.XPATH, "//*[contains(translate(text(),'quickbook','QUICKBOOK'),'Quick Book') or contains(text(),'Quick Book')]")
                )
            )
            log.info("✅ Bus results detected")
            self.screenshot("03_search_results")
            return
        except TimeoutException:
            pass
        log.warning("⚠️  Could not confirm results container — proceeding anyway")
        self.screenshot("03_search_results")

    # ── Step 3b: scrape bus cards ─────────────────────────────────

    def scrape_buses(self) -> List[BusInfo]:
        log.info("🚌 Scraping available buses …")
        buses: List[BusInfo] = []

        # Find all "Quick Book N" buttons — each represents one bus
        book_buttons = self.driver.find_elements(
            By.XPATH, "//*[contains(text(),'Quick Book')]"
        )
        if not book_buttons:
            log.warning("⚠️  No 'Quick Book' buttons found — page source sample:")
            log.warning(self.driver.page_source[:2000])
            return buses

        log.info("   ↳ Found %d bus(es) via Quick Book buttons", len(book_buttons))

        for idx, btn in enumerate(book_buttons):
            b = BusInfo(index=idx)
            b.element_ref = btn  # clicking this button opens the seat map

            # ── Seats from button text: "Quick Book 15" → 15 ─────
            btn_text = btn.text.strip()
            seat_m = re.search(r"\d+", btn_text)
            if seat_m:
                b.available_seats = int(seat_m.group())

            # ── Walk up the DOM to find the card container ────────
            card = btn
            for _ in range(8):
                try:
                    parent = card.find_element(By.XPATH, "..")
                    if len(parent.text.strip()) > len(card.text.strip()) + 20:
                        card = parent
                        break
                    card = parent
                except Exception:
                    break

            text = card.text.strip()

            # ── Bus name / operator ───────────────────────────────
            for sel in ["h2", "h3", "h4", "strong", "b", "[class*='operator' i]", "[class*='name' i]"]:
                try:
                    b.name = card.find_element(By.CSS_SELECTOR, sel).text.strip()
                    if b.name and "Quick Book" not in b.name:
                        break
                except NoSuchElementException:
                    b.name = ""
            if not b.name:
                lines = [l.strip() for l in text.splitlines()
                         if l.strip() and "Quick Book" not in l]
                b.name = lines[0] if lines else f"Bus-{idx+1}"

            # ── Departure time (AM/PM format) ─────────────────────
            time_matches = re.findall(r"\d{1,2}:\d{2}\s*(?:AM|PM)", text, re.IGNORECASE)
            if time_matches:
                b.departure = time_matches[0]

            # ── Fare ─────────────────────────────────────────────
            fare_match = re.search(r"(?:₹|Rs\.?|INR)\s*([\d,]+(?:\.\d+)?)", text)
            if fare_match:
                b.fare = float(fare_match.group(1).replace(",", ""))

            # ── Bus type ─────────────────────────────────────────
            for keyword in ["Non AC", "Non-AC", "AC", "Sleeper", "Deluxe", "Seater", "Volvo"]:
                if keyword.lower() in text.lower():
                    b.bus_type += keyword + " "
            b.bus_type = b.bus_type.strip()

            buses.append(b)
            log.info(
                "   [%d] %s | dep=%s | fare=%.0f | seats=%d | type=%s",
                idx,
                b.name,
                b.departure,
                b.fare,
                b.available_seats,
                b.bus_type,
            )

        return buses


# ─────────────────────────── Step 4: ranking ────────────────────

class BusRanker:
    @staticmethod
    def rank(buses: List[BusInfo]) -> List[BusInfo]:
        if not buses:
            return []

        # Normalise departure times to a priority bonus
        times = [BotBase.time_to_minutes(b.departure) for b in buses]
        max_t = max(times) or 1

        for b, t in zip(buses, times):
            departure_priority = (max_t - t) / max_t * config.WEIGHT_DEPARTURE_BONUS
            b.score = (
                b.fare            * config.WEIGHT_FARE
                + b.available_seats * config.WEIGHT_SEATS
                + departure_priority
            )

        ranked = sorted(buses, key=lambda b: b.score, reverse=True)
        log.info("🏆 Bus ranking:")
        for i, b in enumerate(ranked):
            log.info("   #%d  score=%.2f  %s  fare=%.0f  seats=%d", i+1, b.score, b.name, b.fare, b.available_seats)
        return ranked


# ─────────────────────────── Step 5 & 6: seat selection ─────────

class SeatSelector(BotBase):

    def open_seat_map(self, bus: BusInfo):
        log.info("🗺️  Opening seat map for: %s", bus.name)
        if bus.element_ref:
            btn_text = ""
            try:
                btn_text = bus.element_ref.text.strip()
            except Exception:
                pass
            log.info("   ↳ Clicking: '%s'", btn_text)
            self.safe_click(bus.element_ref)
            time.sleep(4)
            self.screenshot("05_seat_map")

    def select_seats(self) -> list:
        count  = getattr(config, "SEAT_COUNT", 1)
        lo, hi = getattr(config, "SEAT_RANGE", (1, 999))
        log.info("💺 Selecting %d seat(s) from range %d–%d …", count, lo, hi)
        time.sleep(2)
        self._dump_seat_html()

        all_seats = self._get_all_seats()
        if not all_seats:
            log.warning("⚠️  No seat elements detected")
            return []

        available = [s for s in all_seats if self._is_available(s)]
        log.info("   ↳ %d total | %d available", len(all_seats), len(available))

        in_range = [
            s for s in available
            if self._get_seat_label(s).strip().isdigit()
            and lo <= int(self._get_seat_label(s).strip()) <= hi
        ]
        log.info("   ↳ %d seat(s) in range %d–%d", len(in_range), lo, hi)

        if not in_range:
            log.info("   ✅ No seats available in range %d–%d — done with this bus", lo, hi)
            return []
        pool = in_range

        chosen = self._pick_adjacent_seats(pool, count)

        clickers = [
            lambda el: (self.driver.execute_script("arguments[0].focus();", el), time.sleep(0.2), el.send_keys(Keys.SPACE)),
            lambda el: ActionChains(self.driver).move_to_element(el).click().perform(),
            lambda el: self.driver.execute_script("arguments[0].click();", el),
            lambda el: self.driver.execute_script(
                "var el=arguments[0];"
                "['mouseenter','mouseover','mousedown','mouseup','click'].forEach(function(t){"
                "  el.dispatchEvent(new MouseEvent(t,{bubbles:true,cancelable:true,view:window}));"
                "});", el,
            ),
        ]

        selected_ids = []
        for nth, seat in enumerate(chosen, 1):
            seat_id = self._get_seat_label(seat)
            log.info("   ↳ Attempting seat %s (%d/%d) …", seat_id, nth, len(chosen))
            clicked = False
            for attempt, clicker in enumerate(clickers):
                try:
                    self.driver.execute_script("arguments[0].scrollIntoView({block:'center'});", seat)
                    time.sleep(0.5)
                    clicker(seat)
                    time.sleep(1.5)
                    if self._verify_seat_selected(nth):
                        log.info("   ✅ Seat %s selected (strategy %d)", seat_id, attempt + 1)
                        selected_ids.append(seat_id)
                        clicked = True
                        break
                    log.info("   ↳ Strategy %d: not confirmed yet, trying next …", attempt + 1)
                except Exception as e:
                    log.warning("   Strategy %d failed: %s", attempt + 1, e)
            if not clicked:
                log.error("   ❌ All strategies failed for seat %s", seat_id)

        if selected_ids:
            self.screenshot("06_seats_selected")
        else:
            self.screenshot("06_seats_failed")
        return selected_ids

    def _pick_adjacent_seats(self, pool: list, count: int) -> list:
        """
        Return up to `count` seats from the best consecutive run (no padding with
        non-adjacent seats).  Priority: longest run >= count → longest run overall.
        """
        labeled = []
        for seat in pool:
            lbl = self._get_seat_label(seat).strip()
            if lbl.isdigit():
                labeled.append((int(lbl), seat))
        labeled.sort(key=lambda x: x[0])

        if not labeled:
            log.info("   ↳ No numeric seat labels — taking first %d seats", count)
            return pool[:min(count, len(pool))]

        nums     = [x[0] for x in labeled]
        seat_map = {x[0]: x[1] for x in labeled}

        # Build consecutive runs
        runs: list = []
        i = 0
        while i < len(nums):
            j = i
            while j + 1 < len(nums) and nums[j + 1] == nums[j] + 1:
                j += 1
            runs.append(nums[i:j + 1])
            i = j + 1

        # Priority 1: first run that can fill a full block of `count`
        for run in runs:
            if len(run) >= count:
                chosen = run[:count]
                log.info("   ↳ Full consecutive block found: %s", chosen)
                return [seat_map[n] for n in chosen]

        # Priority 2/3: largest run available (already sorted from front)
        best_run = max(runs, key=len)
        chosen = best_run[:count]
        log.info("   ↳ Best consecutive group: %s", chosen)
        return [seat_map[n] for n in chosen]

    def _verify_seat_selected(self, min_count: int = 1) -> bool:
        """Return True if at least min_count seats have the 'selected' class."""
        try:
            selected = self.driver.find_elements(By.CSS_SELECTOR, ".seatCharts-seat.selected")
            if len(selected) >= min_count:
                return True
            checked = self.driver.find_elements(By.CSS_SELECTOR, ".seatCharts-seat[aria-checked='true']")
            if len(checked) >= min_count:
                return True
        except Exception:
            pass
        return False

    def _dump_seat_html(self):
        """Save the seat-map container HTML to a file for debugging."""
        try:
            for sel in [
                "[class*='seat-map']", "[class*='seatmap']", "[class*='seat_map']",
                "[class*='bus-layout']", "[class*='layout']",
            ]:
                els = self.driver.find_elements(By.CSS_SELECTOR, sel)
                if els:
                    html = els[0].get_attribute("outerHTML")
                    Path("seat_map_debug.html").write_text(html, encoding="utf-8")
                    log.info("   ↳ Seat HTML (%d chars) → seat_map_debug.html", len(html))
                    return
            Path("seat_map_debug.html").write_text(self.driver.page_source, encoding="utf-8")
            log.info("   ↳ Full page HTML → seat_map_debug.html")
        except Exception as e:
            log.warning("   ⚠️  Could not dump seat HTML: %s", e)

    def _get_all_seats(self) -> list:
        # ApnaSeat uses the seatCharts jQuery plugin.
        # Available seats have class "available" alongside "seatCharts-seat".
        precise = self.driver.find_elements(
            By.CSS_SELECTOR, ".seatCharts-seat.available"
        )
        if precise:
            return precise

        # Generic fallback
        seat_selectors = [
            "[class*='seat'][class*='available']",
            "[class*='berth'][class*='available']",
            "td[data-seat]",
            "div[data-seat]",
        ]
        for sel in seat_selectors:
            els = self.driver.find_elements(By.CSS_SELECTOR, sel)
            if els:
                return els
        return []

    def _is_available(self, element) -> bool:
        # Only called as extra guard; _get_all_seats() already targets available seats
        classes = (element.get_attribute("class") or "").lower()
        blocked = ["booked", "blocked", "unavailable", "sold", "disabled",
                   "occupied", "reserved"]
        return "available" in classes and not any(k in classes for k in blocked)

    def _get_seat_label(self, element) -> str:
        label = (
            element.get_attribute("data-seat")
            or element.get_attribute("data-seat-number")
            or element.get_attribute("title")
            or element.text.strip()
            or "?"
        )
        return label

    def _pick_by_preference(self, available: list):
        for pref in config.SEAT_PREFERENCE:
            chosen = self._filter_by_pref(available, pref)
            if chosen:
                return chosen[0]
        return available[0] if available else None

    def _filter_by_pref(self, seats: list, pref: str) -> list:
        if pref == "any":
            return seats
        if pref == "window":
            return [s for s in seats if self._is_window(s)]
        if pref == "front":
            return [s for s in seats if self._is_front(s)]
        if pref == "lower":
            return [s for s in seats if self._is_lower(s)]
        return []

    def _is_window(self, element) -> bool:
        classes = (element.get_attribute("class") or "").lower()
        label   = self._get_seat_label(element).lower()
        return "window" in classes or "window" in label or re.search(r"\bA\d|\d A\b", label) is not None

    def _is_front(self, element) -> bool:
        label = self._get_seat_label(element).strip()
        # Only consider purely numeric seat labels (exclude C1, C2, R, etc.)
        if not label.isdigit():
            return False
        return int(label) <= 10

    def _is_lower(self, element) -> bool:
        classes = (element.get_attribute("class") or "").lower()
        label   = self._get_seat_label(element).lower()
        return "lower" in classes or "lower" in label or "l" in classes


# ─────────────────────────── Step 7: passenger form ─────────────

class PassengerForm(BotBase):

    def fill(self, passengers):
        """Fill the booking form for one or more passengers."""
        if not isinstance(passengers, list):
            passengers = [passengers]
        log.info("👤 Filling details for %d passenger(s)", len(passengers))
        time.sleep(2)

        def visible(els):
            return [el for el in els if el.is_displayed()]

        name_els = visible(self.driver.find_elements(
            By.XPATH, "//input[@placeholder='Name *' or @placeholder='Name']"))
        father_els = visible(self.driver.find_elements(
            By.XPATH, "//input[contains(translate(@placeholder,"
            "'abcdefghijklmnopqrstuvwxyz','ABCDEFGHIJKLMNOPQRSTUVWXYZ'),'FATHER')]"))
        id_number_els = visible(self.driver.find_elements(
            By.XPATH, "//input[translate(@placeholder,"
            "'abcdefghijklmnopqrstuvwxyz','ABCDEFGHIJKLMNOPQRSTUVWXYZ')='ID NUMBER *']"))
        residence_els = visible(self.driver.find_elements(
            By.XPATH, "//input[translate(@placeholder,"
            "'abcdefghijklmnopqrstuvwxyz','ABCDEFGHIJKLMNOPQRSTUVWXYZ')='RESIDENCE *']"))
        id_type_els = visible(self.driver.find_elements(
            By.CSS_SELECTOR,
            "select[name*='id' i], select[id*='id_type' i], select[name*='proof' i]"))

        log.info("   ↳ Found: %d name, %d father, %d id_type, %d id_number, %d residence",
                 len(name_els), len(father_els), len(id_type_els),
                 len(id_number_els), len(residence_els))

        pax_count    = len(passengers)
        pax_name_els = name_els[:pax_count]
        contact_name = name_els[pax_count] if len(name_els) > pax_count else None

        for i, pax in enumerate(passengers):
            log.info("   ── Passenger %d: %s | %s | %s",
                     i + 1, pax.name, pax.residence, pax.email)
            if i < len(pax_name_els):
                self._safe_fill(pax_name_els[i], pax.name, f"name[{i}]")
            if i < len(father_els):
                self._safe_fill(father_els[i], pax.father_name or "N/A", f"father[{i}]")
            if i < len(id_type_els):
                self._fill_id_type_el(id_type_els[i], pax.id_type)
            if i < len(id_number_els):
                self._safe_fill(id_number_els[i], pax.id_number, f"id_number[{i}]")
            if i < len(residence_els):
                self._safe_fill(residence_els[i], pax.residence, f"residence[{i}]")

        # ── Contact Detail (uses first passenger) ─────────────────
        first = passengers[0]
        if contact_name:
            self._safe_fill(contact_name, first.name, "contact_name")
        self._fill_field(["email"], first.email)
        self._fill_field(["phone", "mobile", "contact"], first.phone)
        self._fill_boarding_point()
        self._click_confirm_payment()

        self.screenshot("07_passenger_filled")
        log.info("✅ Passenger form filled")

    # ── indexed helpers ───────────────────────────────────────────

    def _safe_fill(self, el, value: str, label: str):
        try:
            self.driver.execute_script("arguments[0].scrollIntoView({block:'center'});", el)
            self.safe_type(el, value)
            log.info("   ↳ %s = '%s'", label, value)
        except Exception as e:
            log.warning("   ⚠️  %s not writable: %s", label, e)

    def _fill_id_type_el(self, sel_el, id_type: str):
        try:
            s = Select(sel_el)
            try:
                s.select_by_visible_text(id_type)
            except Exception:
                s.select_by_value(id_type)
            log.info("   ↳ id_type = '%s'", id_type)
        except Exception as e:
            log.warning("   ⚠️  id_type select error: %s", e)

    # ── placeholder-exact filler ──────────────────────────────────

    def _fill_placeholder_exact(self, placeholder: str, value: str):
        """Fill the first input whose placeholder matches exactly (case-insensitive)."""
        try:
            els = self.driver.find_elements(
                By.XPATH,
                f"//input[translate(@placeholder,'abcdefghijklmnopqrstuvwxyz',"
                f"'ABCDEFGHIJKLMNOPQRSTUVWXYZ')='{placeholder.upper()}']"
            )
            if not els:
                # Fallback: partial match
                els = self.driver.find_elements(
                    By.CSS_SELECTOR,
                    f"input[placeholder*='{placeholder.split()[0]}' i]"
                )
            for el in els:
                if el.is_displayed():
                    try:
                        self.safe_type(el, value)
                        log.info("   ↳ [%s] = '%s'", placeholder, value)
                        return
                    except Exception:
                        pass
            log.warning("   ⚠️  No editable input for placeholder '%s'", placeholder)
        except Exception as e:
            log.warning("   ⚠️  _fill_placeholder_exact error: %s", e)

    # ── boarding point ────────────────────────────────────────────

    def _fill_boarding_point(self):
        """Select the first real option in the Boarding Point dropdown."""
        try:
            # Find by label proximity
            sel_els = self.driver.find_elements(
                By.CSS_SELECTOR,
                "select[name*='boarding' i], select[id*='boarding' i], "
                "select[class*='boarding' i], select[name*='pickup' i]"
            )
            if not sel_els:
                # Broader: find any select near a "Boarding Point" label
                labels = self.driver.find_elements(
                    By.XPATH, "//*[contains(text(),'Boarding Point')]"
                )
                for lbl in labels:
                    try:
                        parent = lbl.find_element(By.XPATH, "..")
                        sel_els = parent.find_elements(By.CSS_SELECTOR, "select")
                        if sel_els:
                            break
                    except Exception:
                        pass

            for sel_el in sel_els:
                try:
                    s = Select(sel_el)
                    # Pick first non-empty option
                    real_opts = [o for o in s.options if o.get_attribute("value")]
                    if real_opts:
                        s.select_by_value(real_opts[0].get_attribute("value"))
                        log.info("   ↳ Boarding Point = '%s'", real_opts[0].text)
                        return
                except Exception as e:
                    log.warning("   ⚠️  Boarding point select error: %s", e)
            log.warning("   ⚠️  Boarding Point dropdown not found")
        except Exception as e:
            log.warning("   ⚠️  _fill_boarding_point error: %s", e)

    # ── confirm payment ───────────────────────────────────────────

    def _click_confirm_payment(self):
        """Click the Confirm / Proceed to Payment button."""
        log.info("💳 Clicking Confirm Payment …")
        time.sleep(1)
        keywords = ["confirm payment", "proceed to payment", "pay now", "confirm booking", "confirm", "proceed", "pay"]
        # Search all visible buttons / links for matching text
        candidates = self.driver.find_elements(By.CSS_SELECTOR, "button, input[type='submit'], a")
        for kw in keywords:
            for el in candidates:
                try:
                    if not el.is_displayed():
                        continue
                    txt = (el.text or el.get_attribute("value") or "").strip().lower()
                    if txt == kw or txt.startswith(kw):
                        log.info("   ↳ Found button: '%s'", el.text.strip() or el.get_attribute("value"))
                        self.safe_click(el)
                        time.sleep(2)
                        self.screenshot("08_payment_page")
                        log.info("   ✅ Confirm Payment clicked")
                        return
                except Exception:
                    pass
        log.warning("   ⚠️  Confirm Payment button not found")
        self.screenshot("08_payment_not_found")

    # ── generic field filler ──────────────────────────────────────

    def _fill_field(self, key_hints: list, value: str):
        for hint in key_hints:
            selectors = [
                f"input[name*='{hint}' i]",
                f"input[id*='{hint}' i]",
                f"input[placeholder*='{hint}' i]",
                f"textarea[name*='{hint}' i]",
            ]
            for sel in selectors:
                try:
                    el = self.driver.find_element(By.CSS_SELECTOR, sel)
                    if el.is_displayed():
                        self.safe_type(el, value)
                        log.info("   ↳ %s = '%s'", hint, value)
                        return
                except NoSuchElementException:
                    pass
        log.warning("   ⚠️  Field '%s' not found", key_hints[0])

    def _fill_gender(self, gender: str):
        # Try radio buttons first
        radios = self.driver.find_elements(
            By.CSS_SELECTOR, "input[type='radio'][value*='male' i], input[type='radio'][value*='female' i]"
        )
        for r in radios:
            val = (r.get_attribute("value") or "").lower()
            if val in gender.lower():
                self.safe_click(r)
                log.info("   ↳ gender radio = '%s'", val)
                return

        # Try <select>
        selects = self.driver.find_elements(
            By.CSS_SELECTOR,
            "select[name*='gender' i], select[id*='gender' i]",
        )
        for sel_el in selects:
            try:
                s = Select(sel_el)
                s.select_by_visible_text(gender)
                log.info("   ↳ gender dropdown = '%s'", gender)
                return
            except Exception:
                pass

        # Fallback: text input
        self._fill_field(["gender"], gender)

    def _fill_id_type(self, id_type: str):
        selects = self.driver.find_elements(
            By.CSS_SELECTOR,
            "select[name*='id' i], select[id*='id_type' i], select[name*='proof' i]",
        )
        for sel_el in selects:
            try:
                s = Select(sel_el)
                s.select_by_visible_text(id_type)
                log.info("   ↳ id_type = '%s'", id_type)
                return
            except Exception:
                try:
                    s.select_by_value(id_type)
                    return
                except Exception:
                    pass

        self._fill_field(["id_type", "proof_type", "document_type"], id_type)


# ─────────────────────────── Excel report ───────────────────────

class BookingWriter:
    """Single append-only workbook: Execution_Summary, Bus_Summary, Passenger_Details."""

    _EXEC_COLS = [
        "Journey Date",
        "Execution Start Time", "Execution End Time", "Total Runtime",
        "Total Buses Processed", "Total Bookings Made", "Total Seats Booked",
        "Failed Booking Attempts",
    ]
    _BUS_COLS = [
        "Execution Start Time", "Bus Route", "Journey Date",
        "Bus Schedule Time", "Bus Operator", "Seats Booked",
        "Booking Count", "Booking Status",
    ]
    _PAX_COLS = [
        "Journey Date", "Bus Schedule Time", "Route Details",
        "Seat Booked", "Passenger Name", "Aadhar Number",
        "Mobile Number", "Father Name", "Email Address",
        "Booking Time", "Booking Status",
    ]

    @staticmethod
    def _filename() -> Path:
        report_dir = Path(getattr(config, "REPORT_DIR", "reports"))
        report_dir.mkdir(exist_ok=True)
        return report_dir / f"Booking_Summary_{datetime.now().strftime('%Y%m%d')}.xlsx"

    @staticmethod
    def _load_or_create(filename: Path):
        from openpyxl import Workbook, load_workbook
        if filename.exists():
            return load_workbook(str(filename))
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Execution_Summary"
        BookingWriter._init_sheet(ws1, BookingWriter._EXEC_COLS)
        BookingWriter._init_sheet(wb.create_sheet("Bus_Summary"), BookingWriter._BUS_COLS)
        BookingWriter._init_sheet(wb.create_sheet("Passenger_Details"), BookingWriter._PAX_COLS)
        return wb

    @staticmethod
    def _init_sheet(ws, cols: list) -> None:
        from openpyxl.styles import Font
        ws.append(cols)
        for cell in ws[1]:
            cell.font = Font(bold=True)
        ws.freeze_panes = "A2"

    @staticmethod
    def _ensure_sheet(wb, name: str, cols: list):
        if name in wb.sheetnames:
            return wb[name]
        ws = wb.create_sheet(name)
        BookingWriter._init_sheet(ws, cols)
        return ws

    @staticmethod
    def _save(wb, filename: Path) -> None:
        from openpyxl.utils import get_column_letter
        for ws in wb.worksheets:
            if ws.max_row < 1:
                continue
            for col_idx, col in enumerate(ws.columns, 1):
                max_len = max(
                    (len(str(cell.value)) for cell in col if cell.value is not None),
                    default=10,
                )
                ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 50)
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions
        wb.save(str(filename))

    @staticmethod
    def append_passengers(
        route: str,
        travel_date: str,
        departure_time: str,
        seat_ids: list,
        passengers: list,
        status: str,
    ) -> Optional[Path]:
        try:
            from openpyxl import Workbook, load_workbook  # noqa: F401
        except ImportError:
            log.warning("⚠️  openpyxl not installed — skipping report")
            return None

        filename = BookingWriter._filename()
        wb = BookingWriter._load_or_create(filename)
        ws = BookingWriter._ensure_sheet(wb, "Passenger_Details", BookingWriter._PAX_COLS)

        booking_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        for i, pax in enumerate(passengers):
            seat = seat_ids[i] if i < len(seat_ids) else "N/A"
            ws.append([
                travel_date, departure_time, route,
                seat, pax.name, pax.id_number, pax.phone,
                pax.father_name, pax.email, booking_time, status,
            ])

        BookingWriter._save(wb, filename)
        log.info("📊 Passenger details appended: %s (%d row(s))", filename, len(passengers))
        return filename

    @staticmethod
    def append_summary(
        summary: dict,
        start_time: datetime,
        end_time: datetime,
        route: str,
        travel_date: str,
    ) -> Optional[Path]:
        try:
            from openpyxl import Workbook, load_workbook  # noqa: F401
        except ImportError:
            log.warning("⚠️  openpyxl not installed — skipping report")
            return None

        filename = BookingWriter._filename()
        wb = BookingWriter._load_or_create(filename)
        runtime = str(end_time - start_time).split(".")[0]
        exec_start = start_time.strftime("%Y-%m-%d %H:%M:%S")

        ws_exec = BookingWriter._ensure_sheet(wb, "Execution_Summary", BookingWriter._EXEC_COLS)
        ws_exec.append([
            travel_date,
            exec_start,
            end_time.strftime("%Y-%m-%d %H:%M:%S"),
            runtime,
            summary["buses_processed"],
            summary["total_bookings"],
            summary["total_seats"],
            summary["failed_attempts"],
        ])

        ws_bus = BookingWriter._ensure_sheet(wb, "Bus_Summary", BookingWriter._BUS_COLS)
        for b in summary["per_bus"]:
            ws_bus.append([
                exec_start,
                route,
                travel_date,
                b["departure"],
                b["name"],
                len(b["seats_booked"]),
                b["bookings"],
                "Success" if b["bookings"] > 0 else "No Seats",
            ])

        BookingWriter._save(wb, filename)
        log.info("📋 Booking summary saved: %s", filename)
        return filename


# ─────────────────────────── cycle helpers ───────────────────────

_WAIT_MINUTES = 30


def _print_cycle_summary(
    summary: dict,
    start_time: datetime,
    end_time: datetime,
    report_path: Optional[Path] = None,
    journey_dates: list = None,
) -> None:
    runtime = str(end_time - start_time).split(".")[0]
    route = f"{config.SOURCE} to {config.DESTINATION}"
    sep = "=" * 50
    date_lines = [f"  {d}" for d in (journey_dates or [])]
    lines = (
        ["", sep, "BOOKING EXECUTION COMPLETED", sep, "Journey Dates Processed:"]
        + date_lines
        + [
            f"Bus Route             : {route}",
            f"Total Buses Processed : {summary.get('buses_processed', 0)}",
            f"Total Bookings Made   : {summary.get('total_bookings', 0)}",
            f"Total Seats Booked    : {summary.get('total_seats', 0)}",
            f"Failed Attempts       : {summary.get('failed_attempts', 0)}",
            f"Report Saved          : {report_path or 'N/A'}",
            f"Started At            : {start_time.strftime('%H:%M:%S')}",
            f"Finished At           : {end_time.strftime('%H:%M:%S')}",
            f"Total Runtime         : {runtime}",
            sep,
            "",
        ]
    )
    print("\n".join(lines), flush=True)


def _countdown(total_seconds: int) -> None:
    log.info("⏳ Waiting %d minutes before next cycle …", total_seconds // 60)
    end = time.time() + total_seconds
    while True:
        remaining = int(end - time.time())
        if remaining <= 0:
            break
        mins = remaining // 60
        if mins % 10 == 0 or mins <= 5:
            log.info("   ⏰ Next cycle in %d minute(s) …", mins)
        time.sleep(min(60, remaining))
    log.info("   ✅ Wait complete — starting next cycle")


# ─────────────────────────── main orchestrator ───────────────────

class ApnaSeatBot:
    """
    Orchestrates all steps:
      1. Open website & search
      2. Process every bus (Change 7)
         – For each bus: exhaust seats 1-23 in groups of up to 4 (Changes 8-10)
      3. Write booking summary (Change 11)
    """

    _MAX_ROUNDS_PER_BUS = 10  # safety cap (23 seats / 4 per round = 6 max)

    def __init__(self):
        self.driver      = BrowserManager.build()
        self.search      = SearchEngine(self.driver)
        self.ranker      = BusRanker()
        self.seats       = SeatSelector(self.driver)
        self.pax_form    = PassengerForm(self.driver)
        self.summary     = {"buses_processed": 0, "total_bookings": 0,
                            "total_seats": 0, "failed_attempts": 0, "per_bus": []}
        self.start_time  = datetime.now()
        self.last_report: Optional[Path] = None

    # ── navigate back to a specific bus seat map ──────────────────

    def _reopen_bus(self, bus: BusInfo, travel_date: str) -> bool:
        """Navigate to home, re-search, and re-open the seat map for `bus`."""
        try:
            log.info("   🔄 Navigating back to search results …")
            self.search.open_website()
            self.search.fill_journey_form(config.SOURCE, config.DESTINATION, travel_date)
            self.search.click_search()
            self.search.wait_for_results()

            book_buttons = self.driver.find_elements(
                By.XPATH, "//*[contains(text(),'Quick Book')]"
            )
            if not book_buttons:
                log.error("   ❌ No Quick Book buttons found after re-search")
                return False

            if bus.index >= len(book_buttons):
                log.error(
                    "   ❌ Bus index %d out of range (%d buttons available)",
                    bus.index, len(book_buttons),
                )
                return False

            bus.element_ref = book_buttons[bus.index]
            log.info(
                "   ↳ Re-opening bus [%d] '%s'", bus.index, bus.element_ref.text.strip()
            )
            self.seats.open_seat_map(bus)
            return True
        except Exception as e:
            log.error("   ❌ _reopen_bus failed: %s", e)
            return False

    # ─────────────────────────────────────────────────────────────

    def run(self, travel_date: str) -> Optional[Path]:
        self.start_time = datetime.now()
        self.summary = {
            "buses_processed": 0,
            "total_bookings":  0,
            "total_seats":     0,
            "failed_attempts": 0,
            "per_bus":         [],
        }
        summary = self.summary

        try:
            # Step 1 — load site and search (used only to discover buses)
            self.search.open_website()
            self.search.fill_journey_form(config.SOURCE, config.DESTINATION, travel_date)
            self.search.click_search()
            self.search.wait_for_results()
            buses = self.search.scrape_buses()

            if not buses:
                log.error("❌ No buses found. Exiting.")
                return False

            ranked = self.ranker.rank(buses)
            route  = f"{config.SOURCE} → {config.DESTINATION}"
            log.info("🚌 Will process all %d bus(es) in ranked order", len(ranked))

            # Step 2 — outer loop: one iteration per bus (Change 7)
            for bus_num, bus in enumerate(ranked, 1):
                log.info("═" * 60)
                log.info(
                    "🚌 Bus %d/%d: %s  dep=%s  score=%.2f",
                    bus_num, len(ranked), bus.name, bus.departure, bus.score,
                )
                bus_stats = {
                    "name":            bus.name,
                    "departure":       bus.departure,
                    "bookings":        0,
                    "seats_booked":    [],
                    "failed_attempts": 0,
                }

                # Inner loop: repeat until seats 1-23 are exhausted (Change 10)
                for booking_round in range(1, self._MAX_ROUNDS_PER_BUS + 1):
                    log.info("   📍 Round %d for %s", booking_round, bus.name)

                    if not self._reopen_bus(bus, travel_date):
                        log.warning("   ⚠️  Cannot re-open bus — stopping rounds")
                        bus_stats["failed_attempts"] += 1
                        summary["failed_attempts"] += 1
                        break

                    seat_ids = self.seats.select_seats()
                    if not seat_ids:
                        log.info(
                            "   ✅ Seats 1-23 exhausted for %s after %d round(s)",
                            bus.name, booking_round - 1,
                        )
                        break

                    n_pax = len(seat_ids)
                    passengers = [generate_passenger() for _ in range(n_pax)]
                    log.info("🎲 Generated %d passenger(s):", n_pax)
                    for i, p in enumerate(passengers):
                        log.info("   [%d] %s | %s | %s", i + 1, p.name, p.residence, p.email)

                    booking_ok = False
                    try:
                        self.pax_form.fill(passengers)
                        booking_ok = True
                        log.info(
                            "   ✅ Round %d booked: seats %s", booking_round, seat_ids
                        )
                    except Exception as e:
                        log.error("   ❌ Passenger form error in round %d: %s", booking_round, e)
                        bus_stats["failed_attempts"] += 1
                        summary["failed_attempts"] += 1

                    status = "Initiated" if booking_ok else "Failed"
                    rp = BookingWriter.append_passengers(
                        route, travel_date, bus.departure, seat_ids, passengers, status
                    )
                    if rp:
                        self.last_report = rp

                    if booking_ok:
                        bus_stats["bookings"] += 1
                        bus_stats["seats_booked"].extend(seat_ids)
                        summary["total_bookings"] += 1
                        summary["total_seats"] += n_pax

                summary["buses_processed"] += 1
                summary["per_bus"].append(bus_stats)
                log.info(
                    "   🏁 Bus %s done — %d booking(s), %d seat(s), %d failure(s)",
                    bus.name, bus_stats["bookings"],
                    len(bus_stats["seats_booked"]), bus_stats["failed_attempts"],
                )

            # Step 3 — summary report (Change 11)
            end_time = datetime.now()
            log.info("═" * 60)
            log.info("🏁 All buses processed!")
            log.info("   Buses processed  : %d", summary["buses_processed"])
            log.info("   Total bookings   : %d", summary["total_bookings"])
            log.info("   Total seats      : %d", summary["total_seats"])
            log.info("   Failed attempts  : %d", summary["failed_attempts"])
            log.info("   Runtime          : %s", str(end_time - self.start_time).split(".")[0])

            rp = BookingWriter.append_summary(summary, self.start_time, end_time, route, travel_date)
            if rp:
                self.last_report = rp

            log.info("⚠️  Review the browser and complete any pending payments manually.")
            return self.last_report

        except Exception as exc:
            log.exception("💥 Unhandled error: %s", exc)
            raise
        finally:
            self.driver.quit()
            log.info("🔒 Browser closed")


# ─────────────────────────── entry point ────────────────────────

def main() -> None:
    cycle = 0
    _SEP = "=" * 50
    route = f"{config.SOURCE} → {config.DESTINATION}"
    log.info("🚀 ApnaSeat Bot started — cycling every %d minutes", _WAIT_MINUTES)
    log.info("   Press Ctrl+C at any time to stop gracefully")

    try:
        while True:
            cycle += 1
            cycle_start = datetime.now()
            dates = get_date_range()
            date_displays = [d.strftime("%Y-%m-%d") for d in dates]

            # ── cycle start banner ──────────────────────────────
            log.info(_SEP)
            log.info("EXECUTION STARTED  (Cycle %d)", cycle)
            log.info(_SEP)
            log.info("Start Time     : %s", cycle_start.strftime("%Y-%m-%d %H:%M:%S"))
            log.info("Route          : %s", route)
            log.info("Journey Dates  :")
            for ds in date_displays:
                log.info("  %s", ds)
            log.info(_SEP)

            cycle_summary = {
                "buses_processed": 0,
                "total_bookings":  0,
                "total_seats":     0,
                "failed_attempts": 0,
                "per_bus":         [],
            }
            last_report: Optional[Path] = None

            for date_obj, date_display in zip(dates, date_displays):
                travel_date = date_obj.strftime("%d-%m-%Y")  # form format

                log.info(_SEP)
                log.info("Journey Date: %s", date_display)
                log.info("Searching buses...")
                log.info(_SEP)

                bot = ApnaSeatBot()
                try:
                    bot.run(travel_date)
                except Exception as exc:
                    log.error("💥 Date %s error: %s", date_display, exc)

                cycle_summary["buses_processed"] += bot.summary.get("buses_processed", 0)
                cycle_summary["total_bookings"]  += bot.summary.get("total_bookings", 0)
                cycle_summary["total_seats"]     += bot.summary.get("total_seats", 0)
                cycle_summary["failed_attempts"] += bot.summary.get("failed_attempts", 0)
                cycle_summary["per_bus"].extend(bot.summary.get("per_bus", []))

                if bot.last_report:
                    last_report = bot.last_report

                log.info("Booking completed for Journey Date: %s", date_display)

            cycle_end = datetime.now()

            # ── cycle end banner ─────────────────────────────────
            log.info(_SEP)
            log.info("EXECUTION COMPLETED  (Cycle %d)", cycle)
            log.info(_SEP)
            log.info("End Time        : %s", cycle_end.strftime("%Y-%m-%d %H:%M:%S"))
            log.info("Runtime         : %s", str(cycle_end - cycle_start).split(".")[0])
            log.info("Journey Dates Processed:")
            for ds in date_displays:
                log.info("  %s", ds)
            log.info("Total Buses Processed   : %d", cycle_summary["buses_processed"])
            log.info("Total Bookings Made     : %d", cycle_summary["total_bookings"])
            log.info("Total Seats Booked      : %d", cycle_summary["total_seats"])
            log.info("Failed Booking Attempts : %d", cycle_summary["failed_attempts"])
            log.info(_SEP)

            _print_cycle_summary(cycle_summary, cycle_start, cycle_end, last_report, date_displays)
            log.info("⏳ Next cycle in %d minutes — press Ctrl+C to stop", _WAIT_MINUTES)
            _countdown(_WAIT_MINUTES * 60)

    except KeyboardInterrupt:
        log.info("⛔ Application stopped by user.")
        print("\nApplication stopped by user.", flush=True)
        sys.exit(0)


if __name__ == "__main__":
    main()
